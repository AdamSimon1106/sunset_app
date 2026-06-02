import time
from pathlib import Path
from datetime import datetime
from functools import lru_cache
from typing import Optional, Tuple, Set

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from tqdm import tqdm

from timezonefinder import TimezoneFinder
import pytz

from astral import LocationInfo
from astral.sun import sun

# ============================================================
# CONFIG
# ============================================================
FLICKR_API_KEY = "0e53a7d7d572227cf99fe4c483e28a13"
FLICKR_API_SECRET = "455e3a2451913503"  # not used for public search

EXCEL_PATH = "sunset_project_v2.xlsx"
SHEET_NAME = 0
TARGET_ROWS = 20000

SUNSET_WINDOW_MINUTES = 60
ELEVATION_WINDOW_METERS = 500

FLICKR_ENDPOINT = "https://api.flickr.com/services/rest/"
ELEVATION_ENDPOINT = "https://api.open-elevation.com/api/v1/lookup"

PER_PAGE = 250
PAGE_SLEEP_SECONDS = 0.15
CHECKPOINT_EVERY_ACCEPTED = 200

FLICKR_TIMEOUT = (10, 30)
ELEVATION_TIMEOUT = (8, 20)

MAX_FLICKR_RETRIES = 4
MAX_ELEVATION_RETRIES = 3

# ============================================================
# GLOBALS
# ============================================================
tf = TimezoneFinder()


def build_session() -> requests.Session:
    session = requests.Session()

    retry = Retry(
        total=3,
        connect=3,
        read=3,
        backoff_factor=1.0,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        raise_on_status=False,
    )

    adapter = HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20)
    session.mount("https://", adapter)
    session.mount("http://", adapter)

    session.headers.update({
        "User-Agent": "sunset-collector/2.0"
    })
    return session


SESSION = build_session()


# ============================================================
# HELPERS
# ============================================================
def resolve_sheet_name(excel_path: str, sheet_name_or_index):
    if isinstance(sheet_name_or_index, str):
        return sheet_name_or_index

    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    names = wb.sheetnames
    idx = int(sheet_name_or_index)

    if idx < 0 or idx >= len(names):
        raise IndexError(f"SHEET_NAME index {idx} out of range. Workbook has sheets: {names}")
    return names[idx]


def pick_best_image_url(p: dict) -> Optional[str]:
    return (
        p.get("url_o")
        or p.get("url_l")
        or p.get("url_c")
        or p.get("url_b")
        or p.get("url_z")
    )


def photo_page_url(owner: str, photo_id: str) -> str:
    return f"https://www.flickr.com/photos/{owner}/{photo_id}"


def _round_coord(x: float, ndigits: int = 5) -> float:
    return round(float(x), ndigits)


def localize_datetaken(
    datetaken_str: str,
    lat: float,
    lon: float
) -> Tuple[Optional[datetime], Optional[datetime], Optional[str]]:
    try:
        local_naive = datetime.strptime(datetaken_str, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None, None, None

    tz_name = tf.timezone_at(lat=lat, lng=lon)
    if not tz_name:
        return None, None, None

    tz = pytz.timezone(tz_name)

    try:
        local_aware = tz.localize(local_naive, is_dst=None)
    except pytz.AmbiguousTimeError:
        local_aware = tz.localize(local_naive, is_dst=True)
    except pytz.NonExistentTimeError:
        local_aware = tz.localize(local_naive, is_dst=True)

    utc_aware = local_aware.astimezone(pytz.utc)
    return local_aware, utc_aware, tz_name


def minutes_from_sunset(
    local_dt_aware: datetime,
    lat: float,
    lon: float,
    tz_name: str
) -> Tuple[float, datetime]:
    tz = pytz.timezone(tz_name)
    loc = LocationInfo(
        name="x",
        region="x",
        timezone=tz_name,
        latitude=lat,
        longitude=lon
    )
    s = sun(loc.observer, date=local_dt_aware.date(), tzinfo=tz)
    sunset_local = s["sunset"]
    diff_minutes = (local_dt_aware - sunset_local).total_seconds() / 60.0
    return diff_minutes, sunset_local


# ============================================================
# NETWORK
# ============================================================
def flickr_call(method: str, **params):
    payload = {
        "method": method,
        "api_key": FLICKR_API_KEY,
        "format": "json",
        "nojsoncallback": 1,
    }
    payload.update(params)

    last_err = None
    for attempt in range(1, MAX_FLICKR_RETRIES + 1):
        try:
            r = SESSION.get(FLICKR_ENDPOINT, params=payload, timeout=FLICKR_TIMEOUT)
            r.raise_for_status()
            data = r.json()

            if data.get("stat") != "ok":
                raise RuntimeError(f"Flickr error: {data}")

            return data
        except Exception as e:
            last_err = e
            sleep_s = 1.2 * attempt
            print(f"[flickr retry] method={method} page={params.get('page')} attempt={attempt} err={e}")
            time.sleep(sleep_s)

    raise RuntimeError(f"Flickr request failed after retries: {last_err}")


@lru_cache(maxsize=100_000)
def get_elevation_m(lat: float, lon: float) -> Optional[float]:
    lat = _round_coord(lat)
    lon = _round_coord(lon)

    last_err = None
    for attempt in range(1, MAX_ELEVATION_RETRIES + 1):
        try:
            r = SESSION.get(
                ELEVATION_ENDPOINT,
                params={"locations": f"{lat},{lon}"},
                timeout=ELEVATION_TIMEOUT,
            )
            r.raise_for_status()
            data = r.json()

            results = data.get("results") or []
            if not results:
                return None

            elev = results[0].get("elevation")
            if elev is None:
                return None

            return float(elev)
        except Exception as e:
            last_err = e
            sleep_s = 0.5 * attempt
            time.sleep(sleep_s)

    print(f"[elevation failed] lat={lat} lon={lon} err={last_err}")
    return None


# ============================================================
# EXCEL / CHECKPOINTS
# ============================================================
def load_schema_df(excel_path: str, sheet_name):
    return pd.read_excel(excel_path, sheet_name=sheet_name)


def load_existing_photo_ids(df_schema: pd.DataFrame) -> Set[str]:
    if "photo_id" not in df_schema.columns:
        return set()

    ids = df_schema["photo_id"].dropna().astype(str).tolist()
    return set(ids)


def align_to_schema(df_new: pd.DataFrame, df_schema: pd.DataFrame) -> pd.DataFrame:
    for col in df_schema.columns:
        if col not in df_new.columns:
            df_new[col] = pd.NA

    schema_cols = list(df_schema.columns)
    extra_cols = [c for c in df_new.columns if c not in schema_cols]
    return df_new[schema_cols + extra_cols]


def write_checkpoint(excel_path: str, sheet_name, df_rows: pd.DataFrame, df_schema: pd.DataFrame):
    df_out = align_to_schema(df_rows.copy(), df_schema)
    out_sheet_name = resolve_sheet_name(excel_path, sheet_name)

    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_out.to_excel(writer, index=False, sheet_name=out_sheet_name)

    print(f"[checkpoint] wrote {len(df_out)} rows to {excel_path}")


# ============================================================
# MAIN FETCH LOGIC
# ============================================================
def fetch_recent_sunsets(df_schema: pd.DataFrame) -> pd.DataFrame:
    existing_ids = load_existing_photo_ids(df_schema)

    rows = []
    seen = set(existing_ids)

    page = 1

    scanned = 0
    rejected_basic = 0
    rejected_tz = 0
    rejected_sunset = 0
    rejected_elev = 0
    elev_fail = 0

    extras = ",".join([
        "date_taken",
        "geo",
        "url_o",
        "url_l",
        "url_c",
        "url_b",
        "url_z",
    ])

    with tqdm(total=TARGET_ROWS, desc="Collecting valid sunset-window + sea-level photos") as pbar:
        while len(rows) < TARGET_ROWS:
            page_start = time.time()

            data = flickr_call(
                "flickr.photos.search",
                text="sunset",
                tags="sunset",
                tag_mode="any",
                sort="date-posted-desc",
                content_type=1,
                media="photos",
                safe_search=1,
                has_geo=1,
                extras=extras,
                per_page=PER_PAGE,
                page=page,
            )

            photos = data.get("photos", {}).get("photo", [])
            if not photos:
                print("[done] no more Flickr photos returned")
                break

            page_accept_before = len(rows)

            for p in photos:
                if len(rows) >= TARGET_ROWS:
                    break

                pid = p.get("id")
                if not pid:
                    rejected_basic += 1
                    continue

                pid = str(pid)
                if pid in seen:
                    continue
                seen.add(pid)

                dt_str = p.get("datetaken")
                img_url = pick_best_image_url(p)

                try:
                    lat = float(p.get("latitude"))
                    lon = float(p.get("longitude"))
                except (TypeError, ValueError):
                    rejected_basic += 1
                    continue

                if (lat == 0.0 and lon == 0.0) or (not dt_str) or (not img_url):
                    rejected_basic += 1
                    continue

                scanned += 1

                local_dt, utc_dt, tz_name = localize_datetaken(dt_str, lat, lon)
                if not local_dt or not utc_dt or not tz_name:
                    rejected_tz += 1
                    continue

                try:
                    mins_from, _sunset_local = minutes_from_sunset(local_dt, lat, lon, tz_name)
                except Exception:
                    rejected_sunset += 1
                    continue

                if abs(mins_from) > SUNSET_WINDOW_MINUTES:
                    rejected_sunset += 1
                    continue

                elev = get_elevation_m(lat, lon)
                if elev is None:
                    elev_fail += 1
                    continue

                if abs(elev) > ELEVATION_WINDOW_METERS:
                    rejected_elev += 1
                    continue

                owner = p.get("owner")
                flickr_url = photo_page_url(owner, pid) if owner else None

                rows.append({
                    "datetime": utc_dt.isoformat(),
                    "lat": lat,
                    "lon": lon,
                    "elevation_m": float(elev),
                    "sunset_direction": pd.NA,
                    "cloud_cover_total": pd.NA,
                    "cloud_cover_low": pd.NA,
                    "cloud_cover_mid": pd.NA,
                    "cloud_cover_high": pd.NA,
                    "cloud_type": pd.NA,
                    "relative_humidity": pd.NA,
                    "absolute_humidity": pd.NA,
                    "dew_point": pd.NA,
                    "temperature": pd.NA,
                    "pm25": pd.NA,
                    "pm10": pd.NA,
                    "dust": pd.NA,
                    "smoke": pd.NA,
                    "aerosol_optical_depth": pd.NA,
                    "visibility": pd.NA,
                    "wind_speed": pd.NA,
                    "wind_direction": pd.NA,
                    "rain_last_6h": pd.NA,
                    "rain_last_12h": pd.NA,
                    "pressure": pd.NA,
                    "pressure_6h_before": pd.NA,
                    "pressure_trend": pd.NA,
                    "solar_elevation": pd.NA,
                    "solar_azimuth": pd.NA,
                    "minutes_from_sunset": float(mins_from),
                    "photo_id": pid,
                    "flickr_url": flickr_url,
                    "beauty_score": pd.NA,
                    "cloud_cover_total_sunset_dir": pd.NA,
                    "cloud_cover_low_sunset_dir": pd.NA,
                    "cloud_cover_mid_sunset_dir": pd.NA,
                    "cloud_cover_high_sunset_dir": pd.NA,
                    "cloud_type_sunset_dir": pd.NA,
                    "aerosol_optical_depth_sunset_dir": pd.NA,
                    "pm25_sunset_dir": pd.NA,
                    "pm10_sunset_dir": pd.NA,
                    "dust_sunset_dir": pd.NA,
                    "smoke_sunset_dir": pd.NA,
                    "visibility_sunset_dir": pd.NA,
                    "rain_last_6h_sunset_dir": pd.NA,
                    "rain_last_12h_sunset_dir": pd.NA,
                })

                pbar.update(1)

                if len(rows) % CHECKPOINT_EVERY_ACCEPTED == 0:
                    write_checkpoint(EXCEL_PATH, SHEET_NAME, pd.DataFrame(rows), df_schema)

                if scanned % 50 == 0:
                    print(
                        f"[status] scanned={scanned} accepted={len(rows)} "
                        f"rej_basic={rejected_basic} rej_tz={rejected_tz} "
                        f"rej_sunset={rejected_sunset} elev_fail={elev_fail} "
                        f"rej_elev={rejected_elev} page={page}"
                    )

            page_elapsed = time.time() - page_start
            page_added = len(rows) - page_accept_before

            print(
                f"[page done] page={page} photos={len(photos)} "
                f"accepted_this_page={page_added} total_accepted={len(rows)} "
                f"elapsed={page_elapsed:.2f}s"
            )

            page += 1
            time.sleep(PAGE_SLEEP_SECONDS)

    print(
        f"[done] scanned={scanned} accepted={len(rows)} "
        f"rej_basic={rejected_basic} rej_tz={rejected_tz} "
        f"rej_sunset={rejected_sunset} elev_fail={elev_fail} "
        f"rej_elev={rejected_elev}"
    )

    return pd.DataFrame(rows)


# ============================================================
# ENTRYPOINT
# ============================================================
def main():
    if not Path(EXCEL_PATH).exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

    df_schema = load_schema_df(EXCEL_PATH, SHEET_NAME)
    df_new = fetch_recent_sunsets(df_schema)

    df_out = align_to_schema(df_new, df_schema)
    out_sheet_name = resolve_sheet_name(EXCEL_PATH, SHEET_NAME)

    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_out.to_excel(writer, index=False, sheet_name=out_sheet_name)

    print(
        f"Wrote {len(df_out)} rows "
        f"(sunset-window, elevation ±{ELEVATION_WINDOW_METERS}m, UTC) into {EXCEL_PATH}"
    )


if __name__ == "__main__":
    main()